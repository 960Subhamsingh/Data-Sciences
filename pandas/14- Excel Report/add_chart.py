from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


# Read workbook and select sheet
wb = load_workbook('D:/Project/Data Sciences/pandas/14- Excel Report/pivot_table.xlsx')
sheet = wb['Report']

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(f'Min Columns: {min_column}')
print(f'Max Columns: {max_column}')
print(f'Min Rows: {min_row}')
print(f'Max Rows: {max_row}')

# Adding Excel charts through Python

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)  # including headers

categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)  # not including headers

# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style

# Save workbook
wb.save('D:/Project/Data Sciences/pandas/14- Excel Report/barchart.xlsx')

 