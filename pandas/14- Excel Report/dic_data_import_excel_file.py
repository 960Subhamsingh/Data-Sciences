from openpyxl import Workbook, load_workbook
from openpyxl.utils import  get_column_letter
from openpyxl.styles import Font

data = {
	"subham kumar": {
		"math": 65,
		"science": 78,
		"english": 98,
		"Hindi": 89
	},
	"Monu kuamr": {
		"math": 55,
		"science": 72,
		"english": 87,
		"Hindi": 95
	},
	"Ram ": {
		"math": 100,
		"science": 45,
		"english": 75,
		"Hindi": 92
	},
	"Sonu Kumar": {
		"math": 30,
		"science": 25,
		"english": 45,
		"Hindi": 100
	},
	"Shivam": {
		"math": 100,
		"science": 100,
		"english": 100,
		"Hindi": 60
	},
    "sourav": {
        "math": 35,
        "science":34,
        "english":34,
        "hindi":45
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['subham kumar'].keys())
ws.append(headings)


for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['subham kumar']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("D:/Project/Data Sciences/pandas/14- Excel Report/NewGrades.xlsx")